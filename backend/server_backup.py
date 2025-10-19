from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.responses import StreamingResponse
import aiomysql
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import json
import openpyxl
import io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MySQL connection pool
pool = None

async def get_db_pool():
    global pool
    if pool is None:
        pool = await aiomysql.create_pool(
            host=os.environ.get('MYSQL_HOST', 'localhost'),
            port=int(os.environ.get('MYSQL_PORT', '3306')),
            user=os.environ.get('MYSQL_USER', 'root'),
            password=os.environ.get('MYSQL_PASSWORD', ''),
            db=os.environ.get('MYSQL_DATABASE', 'avlv_quotes'),
            autocommit=True,
            minsize=1,
            maxsize=10
        )
    return pool

async def get_db():
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        async with conn.cursor(aiomysql.DictCursor) as cur:
            yield cur

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# ============ Models ============

class UserCreate(BaseModel):
    username: str
    password: str
    role: str = 'estimator'
    department_id: Optional[int] = None

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: int
    username: str
    role: str
    department_id: Optional[int]
    created_at: str

class DepartmentCreate(BaseModel):
    name: str

class DepartmentResponse(BaseModel):
    id: int
    name: str
    created_at: str

class QuoteCreate(BaseModel):
    name: str
    client_name: str
    department_id: int
    description: Optional[str] = None

class QuoteUpdate(BaseModel):
    name: Optional[str] = None
    client_name: Optional[str] = None
    department_id: Optional[int] = None
    description: Optional[str] = None
    status: Optional[str] = None

class QuoteResponse(BaseModel):
    id: int
    name: str
    client_name: str
    department_id: int
    department_name: Optional[str]
    description: Optional[str]
    status: str
    version: int
    created_by: int
    created_by_username: Optional[str]
    created_at: str
    updated_at: str

class RoomCreate(BaseModel):
    quote_id: int
    name: str
    system_type: str

class RoomResponse(BaseModel):
    id: int
    quote_id: int
    name: str
    system_type: str
    created_at: str

class EquipmentCreate(BaseModel):
    room_id: int
    item_name: str
    description: Optional[str] = None
    quantity: int
    unit_price: float
    vendor: Optional[str] = None

class EquipmentResponse(BaseModel):
    id: int
    room_id: int
    item_name: str
    description: Optional[str]
    quantity: int
    unit_price: float
    vendor: Optional[str]
    total_price: float
    created_at: str

class LaborCreate(BaseModel):
    room_id: int
    role_name: str
    rate: float
    hours: float
    department_id: Optional[int]

class LaborResponse(BaseModel):
    id: int
    room_id: int
    role_name: str
    rate: float
    hours: float
    department_id: Optional[int]
    total_cost: float
    created_at: str

class ServiceCreate(BaseModel):
    room_id: int
    service_name: str
    cost: float
    department_id: Optional[int]
    description: Optional[str]

class ServiceResponse(BaseModel):
    id: int
    room_id: int
    service_name: str
    cost: float
    department_id: Optional[int]
    description: Optional[str]
    created_at: str

class VendorPriceCreate(BaseModel):
    item_name: str
    price: float
    description: Optional[str]
    vendor: str
    department_id: Optional[int]

class VendorPriceResponse(BaseModel):
    id: int
    item_name: str
    price: float
    description: Optional[str]
    vendor: str
    department_id: Optional[int]
    imported_at: str

class MetricCreate(BaseModel):
    metric_name: str
    metric_type: str
    config: Dict[str, Any]

class MetricResponse(BaseModel):
    id: int
    user_id: int
    metric_name: str
    metric_type: str
    config: Dict[str, Any]
    created_at: str

class ApprovalCreate(BaseModel):
    quote_id: int
    approver_id: int
    status: str = 'pending'
    notes: Optional[str]

class ApprovalResponse(BaseModel):
    id: int
    quote_id: int
    approver_id: int
    approver_username: Optional[str]
    status: str
    notes: Optional[str]
    created_at: str
    updated_at: str

# ============ Auth Helper Functions ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: int, username: str, role: str) -> str:
    payload = {
        'user_id': user_id,
        'username': username,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(user = Depends(get_current_user)):
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

# ============ Database Initialization ============

async def init_db():
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            # Users table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    username VARCHAR(255) UNIQUE NOT NULL,
                    password_hash VARCHAR(255) NOT NULL,
                    role VARCHAR(50) NOT NULL,
                    department_id INT,
                    azure_enabled BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_username (username),
                    INDEX idx_role (role)
                )
            """)

            # Departments table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS departments (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) UNIQUE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Quotes table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS quotes (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    client_name VARCHAR(255) NOT NULL,
                    department_id INT NOT NULL,
                    description TEXT,
                    status VARCHAR(50) DEFAULT 'draft',
                    version INT DEFAULT 1,
                    created_by INT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    FOREIGN KEY (created_by) REFERENCES users(id),
                    INDEX idx_status (status),
                    INDEX idx_department (department_id),
                    INDEX idx_created_by (created_by)
                )
            """)

            # Quote versions (history)
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS quote_versions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    quote_id INT NOT NULL,
                    version INT NOT NULL,
                    data JSON NOT NULL,
                    changed_by INT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE,
                    FOREIGN KEY (changed_by) REFERENCES users(id),
                    INDEX idx_quote_version (quote_id, version)
                )
            """)

            # Rooms table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS rooms (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    quote_id INT NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    system_type VARCHAR(255) NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE,
                    INDEX idx_quote (quote_id)
                )
            """)

            # Equipment table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS equipment (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    room_id INT NOT NULL,
                    item_name VARCHAR(255) NOT NULL,
                    description TEXT,
                    quantity INT NOT NULL,
                    unit_price DECIMAL(10, 2) NOT NULL,
                    vendor VARCHAR(255),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
                    INDEX idx_room (room_id)
                )
            """)

            # Labor table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS labor (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    room_id INT NOT NULL,
                    role_name VARCHAR(255) NOT NULL,
                    rate DECIMAL(10, 2) NOT NULL,
                    hours DECIMAL(10, 2) NOT NULL,
                    department_id INT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    INDEX idx_room (room_id)
                )
            """)

            # Third-party services table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS services (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    room_id INT NOT NULL,
                    service_name VARCHAR(255) NOT NULL,
                    cost DECIMAL(10, 2) NOT NULL,
                    department_id INT,
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    INDEX idx_room (room_id)
                )
            """)

            # Vendor prices table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS vendor_prices (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    item_name VARCHAR(255) NOT NULL,
                    price DECIMAL(10, 2) NOT NULL,
                    description TEXT,
                    vendor VARCHAR(255) NOT NULL,
                    department_id INT,
                    imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    INDEX idx_vendor (vendor),
                    INDEX idx_item (item_name)
                )
            """)

            # Metrics table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS metrics (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    metric_name VARCHAR(255) NOT NULL,
                    metric_type VARCHAR(100) NOT NULL,
                    config JSON NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                    INDEX idx_user (user_id)
                )
            """)

            # Approvals table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS approvals (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    quote_id INT NOT NULL,
                    approver_id INT NOT NULL,
                    status VARCHAR(50) DEFAULT 'pending',
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE,
                    FOREIGN KEY (approver_id) REFERENCES users(id),
                    INDEX idx_quote (quote_id),
                    INDEX idx_status (status)
                )
            """)

            # Create default admin user if not exists
            await cur.execute("SELECT id FROM users WHERE username = 'admin'")
            if not await cur.fetchone():
                admin_password = hash_password('admin123')
                await cur.execute(
                    "INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s)",
                    ('admin', admin_password, 'admin')
                )

# ============ Auth Routes ============

@api_router.post("/auth/register")
async def register(user_data: UserCreate, current_user = Depends(require_admin), cur = Depends(get_db)):
    """Admin only - register new users"""
    password_hash = hash_password(user_data.password)
    try:
        await cur.execute(
            "INSERT INTO users (username, password_hash, role, department_id) VALUES (%s, %s, %s, %s)",
            (user_data.username, password_hash, user_data.role, user_data.department_id)
        )
        return {"message": "User created successfully"}
    except aiomysql.IntegrityError:
        raise HTTPException(status_code=400, detail="Username already exists")

@api_router.post("/auth/login")
async def login(credentials: UserLogin, cur = Depends(get_db)):
    await cur.execute("SELECT * FROM users WHERE username = %s", (credentials.username,))
    user = await cur.fetchone()
    
    if not user or not verify_password(credentials.password, user['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user['id'], user['username'], user['role'])
    return {
        "token": token,
        "user": {
            "id": user['id'],
            "username": user['username'],
            "role": user['role'],
            "department_id": user['department_id']
        }
    }

@api_router.get("/auth/me")
async def get_me(user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT id, username, role, department_id, created_at FROM users WHERE id = %s", (user['user_id'],))
    user_data = await cur.fetchone()
    if not user_data:
        raise HTTPException(status_code=404, detail="User not found")
    return user_data

# ============ Department Routes ============

@api_router.post("/departments")
async def create_department(dept: DepartmentCreate, user = Depends(require_admin), cur = Depends(get_db)):
    try:
        await cur.execute("INSERT INTO departments (name) VALUES (%s)", (dept.name,))
        return {"message": "Department created successfully"}
    except aiomysql.IntegrityError:
        raise HTTPException(status_code=400, detail="Department already exists")

@api_router.get("/departments")
async def get_departments(user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM departments ORDER BY name")
    return await cur.fetchall()

@api_router.put("/departments/{dept_id}")
async def update_department(dept_id: int, dept: DepartmentCreate, user = Depends(require_admin), cur = Depends(get_db)):
    await cur.execute("UPDATE departments SET name = %s WHERE id = %s", (dept.name, dept_id))
    return {"message": "Department updated successfully"}

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: int, user = Depends(require_admin), cur = Depends(get_db)):
    await cur.execute("DELETE FROM departments WHERE id = %s", (dept_id,))
    return {"message": "Department deleted successfully"}

# ============ Quote Routes ============

@api_router.post("/quotes")
async def create_quote(quote: QuoteCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO quotes (name, client_name, department_id, description, created_by) VALUES (%s, %s, %s, %s, %s)",
        (quote.name, quote.client_name, quote.department_id, quote.description, user['user_id'])
    )
    quote_id = cur.lastrowid
    return {"id": quote_id, "message": "Quote created successfully"}

@api_router.get("/quotes")
async def get_quotes(user = Depends(get_current_user), cur = Depends(get_db)):
    # Filter by department if user is not admin
    if user['role'] == 'admin':
        await cur.execute("""
            SELECT q.*, d.name as department_name, u.username as created_by_username
            FROM quotes q
            LEFT JOIN departments d ON q.department_id = d.id
            LEFT JOIN users u ON q.created_by = u.id
            ORDER BY q.updated_at DESC
        """)
    else:
        await cur.execute("""
            SELECT q.*, d.name as department_name, u.username as created_by_username
            FROM quotes q
            LEFT JOIN departments d ON q.department_id = d.id
            LEFT JOIN users u ON q.created_by = u.id
            WHERE q.department_id = (SELECT department_id FROM users WHERE id = %s)
            ORDER BY q.updated_at DESC
        """, (user['user_id'],))
    
    return await cur.fetchall()

@api_router.get("/quotes/{quote_id}")
async def get_quote(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("""
        SELECT q.*, d.name as department_name, u.username as created_by_username
        FROM quotes q
        LEFT JOIN departments d ON q.department_id = d.id
        LEFT JOIN users u ON q.created_by = u.id
        WHERE q.id = %s
    """, (quote_id,))
    quote = await cur.fetchone()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    return quote

@api_router.put("/quotes/{quote_id}")
async def update_quote(quote_id: int, quote_data: QuoteUpdate, user = Depends(get_current_user), cur = Depends(get_db)):
    # Get current version
    await cur.execute("SELECT version, department_id FROM quotes WHERE id = %s", (quote_id,))
    current = await cur.fetchone()
    if not current:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Save current state to version history
    await cur.execute("SELECT * FROM quotes WHERE id = %s", (quote_id,))
    quote_snapshot = await cur.fetchone()
    await cur.execute(
        "INSERT INTO quote_versions (quote_id, version, data, changed_by) VALUES (%s, %s, %s, %s)",
        (quote_id, current['version'], json.dumps(quote_snapshot, default=str), user['user_id'])
    )
    
    # Update quote and increment version
    update_fields = []
    values = []
    for field, value in quote_data.model_dump(exclude_unset=True).items():
        update_fields.append(f"{field} = %s")
        values.append(value)
    
    if update_fields:
        update_fields.append("version = version + 1")
        values.append(quote_id)
        await cur.execute(
            f"UPDATE quotes SET {', '.join(update_fields)} WHERE id = %s",
            tuple(values)
        )
    
    return {"message": "Quote updated successfully"}

@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM quotes WHERE id = %s", (quote_id,))
    return {"message": "Quote deleted successfully"}

@api_router.get("/quotes/{quote_id}/versions")
async def get_quote_versions(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("""
        SELECT qv.*, u.username as changed_by_username
        FROM quote_versions qv
        LEFT JOIN users u ON qv.changed_by = u.id
        WHERE qv.quote_id = %s
        ORDER BY qv.version DESC
    """, (quote_id,))
    return await cur.fetchall()

# ============ Room Routes ============

@api_router.post("/rooms")
async def create_room(room: RoomCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO rooms (quote_id, name, system_type) VALUES (%s, %s, %s)",
        (room.quote_id, room.name, room.system_type)
    )
    return {"id": cur.lastrowid, "message": "Room created successfully"}

@api_router.get("/rooms/quote/{quote_id}")
async def get_rooms_by_quote(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM rooms WHERE quote_id = %s", (quote_id,))
    return await cur.fetchall()

@api_router.delete("/rooms/{room_id}")
async def delete_room(room_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM rooms WHERE id = %s", (room_id,))
    return {"message": "Room deleted successfully"}

# ============ Equipment Routes ============

@api_router.post("/equipment")
async def create_equipment(equip: EquipmentCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO equipment (room_id, item_name, description, quantity, unit_price, vendor) VALUES (%s, %s, %s, %s, %s, %s)",
        (equip.room_id, equip.item_name, equip.description, equip.quantity, equip.unit_price, equip.vendor)
    )
    return {"id": cur.lastrowid, "message": "Equipment created successfully"}

@api_router.get("/equipment/room/{room_id}")
async def get_equipment_by_room(room_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT *, (quantity * unit_price) as total_price FROM equipment WHERE room_id = %s", (room_id,))
    return await cur.fetchall()

@api_router.delete("/equipment/{equip_id}")
async def delete_equipment(equip_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM equipment WHERE id = %s", (equip_id,))
    return {"message": "Equipment deleted successfully"}

# ============ Labor Routes ============

@api_router.post("/labor")
async def create_labor(labor: LaborCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO labor (room_id, role_name, rate, hours, department_id) VALUES (%s, %s, %s, %s, %s)",
        (labor.room_id, labor.role_name, labor.rate, labor.hours, labor.department_id)
    )
    return {"id": cur.lastrowid, "message": "Labor created successfully"}

@api_router.get("/labor/room/{room_id}")
async def get_labor_by_room(room_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT *, (rate * hours) as total_cost FROM labor WHERE room_id = %s", (room_id,))
    return await cur.fetchall()

@api_router.delete("/labor/{labor_id}")
async def delete_labor(labor_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM labor WHERE id = %s", (labor_id,))
    return {"message": "Labor deleted successfully"}

# ============ Service Routes ============

@api_router.post("/services")
async def create_service(service: ServiceCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO services (room_id, service_name, cost, department_id, description) VALUES (%s, %s, %s, %s, %s)",
        (service.room_id, service.service_name, service.cost, service.department_id, service.description)
    )
    return {"id": cur.lastrowid, "message": "Service created successfully"}

@api_router.get("/services/room/{room_id}")
async def get_services_by_room(room_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM services WHERE room_id = %s", (room_id,))
    return await cur.fetchall()

@api_router.delete("/services/{service_id}")
async def delete_service(service_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM services WHERE id = %s", (service_id,))
    return {"message": "Service deleted successfully"}

# ============ Vendor Price Routes ============

@api_router.post("/vendor-prices/import")
async def import_vendor_prices(file: UploadFile = File(...), user = Depends(get_current_user), cur = Depends(get_db)):
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active
    
    # Get headers
    headers = [cell.value for cell in sheet[1]]
    return {"headers": headers, "row_count": sheet.max_row - 1}

@api_router.post("/vendor-prices/import-mapped")
async def import_vendor_prices_mapped(
    file: UploadFile = File(...),
    mapping: str = Form(...),
    vendor: str = Form(...),
    department_id: Optional[int] = Form(None),
    user = Depends(get_current_user),
    cur = Depends(get_db)
):
    mapping_dict = json.loads(mapping)
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active
    
    headers = [cell.value for cell in sheet[1]]
    imported_count = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item_name = row[headers.index(mapping_dict['item_name'])] if mapping_dict.get('item_name') else None
        price = row[headers.index(mapping_dict['price'])] if mapping_dict.get('price') else None
        description = row[headers.index(mapping_dict['description'])] if mapping_dict.get('description') and mapping_dict['description'] in headers else None
        
        if item_name and price:
            try:
                await cur.execute(
                    "INSERT INTO vendor_prices (item_name, price, description, vendor, department_id) VALUES (%s, %s, %s, %s, %s)",
                    (str(item_name), float(price), str(description) if description else None, vendor, department_id)
                )
                imported_count += 1
            except Exception as e:
                logging.error(f"Error importing row: {e}")
    
    return {"message": f"Imported {imported_count} items successfully"}

@api_router.get("/vendor-prices")
async def get_vendor_prices(user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM vendor_prices ORDER BY imported_at DESC")
    return await cur.fetchall()

@api_router.get("/vendor-prices/search")
async def search_vendor_prices(q: str, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "SELECT * FROM vendor_prices WHERE item_name LIKE %s OR description LIKE %s LIMIT 50",
        (f"%{q}%", f"%{q}%")
    )
    return await cur.fetchall()

# ============ Dashboard & Metrics Routes ============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user = Depends(get_current_user), cur = Depends(get_db)):
    stats = {}
    
    # Total quotes by department
    if user['role'] == 'admin':
        await cur.execute("""
            SELECT d.name, COUNT(q.id) as count
            FROM departments d
            LEFT JOIN quotes q ON d.id = q.department_id
            GROUP BY d.id, d.name
        """)
    else:
        await cur.execute("""
            SELECT d.name, COUNT(q.id) as count
            FROM departments d
            LEFT JOIN quotes q ON d.id = q.department_id
            WHERE d.id = (SELECT department_id FROM users WHERE id = %s)
            GROUP BY d.id, d.name
        """, (user['user_id'],))
    stats['quotes_by_department'] = await cur.fetchall()
    
    # Quote status breakdown
    dept_filter = "" if user['role'] == 'admin' else f"WHERE department_id = (SELECT department_id FROM users WHERE id = {user['user_id']})"
    await cur.execute(f"""
        SELECT status, COUNT(*) as count
        FROM quotes
        {dept_filter}
        GROUP BY status
    """)
    stats['quotes_by_status'] = await cur.fetchall()
    
    # Recent quotes (90 days)
    await cur.execute(f"""
        SELECT DATE(created_at) as date, COUNT(*) as count
        FROM quotes
        WHERE created_at >= DATE_SUB(NOW(), INTERVAL 90 DAY)
        {dept_filter}
        GROUP BY DATE(created_at)
        ORDER BY date
    """)
    stats['recent_quotes'] = await cur.fetchall()
    
    # Revenue by department (estimated)
    await cur.execute(f"""
        SELECT d.name, 
               SUM(e.quantity * e.unit_price) as equipment_total,
               SUM(l.rate * l.hours) as labor_total,
               SUM(s.cost) as services_total
        FROM departments d
        LEFT JOIN quotes q ON d.id = q.department_id
        LEFT JOIN rooms r ON q.id = r.quote_id
        LEFT JOIN equipment e ON r.id = e.room_id
        LEFT JOIN labor l ON r.id = l.room_id
        LEFT JOIN services s ON r.id = s.room_id
        {"" if user['role'] == 'admin' else f"WHERE d.id = (SELECT department_id FROM users WHERE id = {user['user_id']})"}
        GROUP BY d.id, d.name
    """)
    stats['revenue_by_department'] = await cur.fetchall()
    
    return stats

@api_router.post("/metrics")
async def create_metric(metric: MetricCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO metrics (user_id, metric_name, metric_type, config) VALUES (%s, %s, %s, %s)",
        (user['user_id'], metric.metric_name, metric.metric_type, json.dumps(metric.config))
    )
    return {"id": cur.lastrowid, "message": "Metric created successfully"}

@api_router.get("/metrics")
async def get_metrics(user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM metrics WHERE user_id = %s", (user['user_id'],))
    metrics = await cur.fetchall()
    for m in metrics:
        if isinstance(m['config'], str):
            m['config'] = json.loads(m['config'])
    return metrics

@api_router.delete("/metrics/{metric_id}")
async def delete_metric(metric_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM metrics WHERE id = %s AND user_id = %s", (metric_id, user['user_id']))
    return {"message": "Metric deleted successfully"}

# ============ Approval Routes ============

@api_router.post("/approvals")
async def create_approval(approval: ApprovalCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO approvals (quote_id, approver_id, status, notes) VALUES (%s, %s, %s, %s)",
        (approval.quote_id, approval.approver_id, approval.status, approval.notes)
    )
    return {"id": cur.lastrowid, "message": "Approval request created"}

@api_router.get("/approvals/quote/{quote_id}")
async def get_approvals_by_quote(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("""
        SELECT a.*, u.username as approver_username
        FROM approvals a
        LEFT JOIN users u ON a.approver_id = u.id
        WHERE a.quote_id = %s
        ORDER BY a.created_at DESC
    """, (quote_id,))
    return await cur.fetchall()

@api_router.put("/approvals/{approval_id}")
async def update_approval(approval_id: int, status: str, notes: Optional[str] = None, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "UPDATE approvals SET status = %s, notes = %s WHERE id = %s",
        (status, notes, approval_id)
    )
    return {"message": "Approval updated successfully"}

# ============ PDF Generation ============

@api_router.get("/quotes/{quote_id}/pdf")
async def generate_quote_pdf(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    # Get quote details
    await cur.execute("""
        SELECT q.*, d.name as department_name, u.username as created_by_username
        FROM quotes q
        LEFT JOIN departments d ON q.department_id = d.id
        LEFT JOIN users u ON q.created_by = u.id
        WHERE q.id = %s
    """, (quote_id,))
    quote = await cur.fetchone()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Get rooms with equipment, labor, services
    await cur.execute("SELECT * FROM rooms WHERE quote_id = %s", (quote_id,))
    rooms = await cur.fetchall()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"Quote: {quote['name']}", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Quote info
    info_data = [
        ['Client:', quote['client_name']],
        ['Department:', quote['department_name']],
        ['Status:', quote['status'].upper()],
        ['Version:', str(quote['version'])],
        ['Created By:', quote['created_by_username']],
        ['Date:', str(quote['created_at'])],
    ]
    if quote['description']:
        info_data.append(['Description:', quote['description']])
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Room breakdown
    total_equipment = 0
    total_labor = 0
    total_services = 0
    
    for room in rooms:
        elements.append(Paragraph(f"Room: {room['name']} ({room['system_type']})", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Equipment
        await cur.execute("SELECT * FROM equipment WHERE room_id = %s", (room['id'],))
        equipment = await cur.fetchall()
        if equipment:
            elements.append(Paragraph("Equipment:", styles['Heading3']))
            equip_data = [['Item', 'Qty', 'Unit Price', 'Total']]
            for eq in equipment:
                total = float(eq['quantity']) * float(eq['unit_price'])
                total_equipment += total
                equip_data.append([
                    eq['item_name'],
                    str(eq['quantity']),
                    f"${eq['unit_price']:.2f}",
                    f"${total:.2f}"
                ])
            
            equip_table = Table(equip_data, colWidths=[3*inch, 0.8*inch, 1*inch, 1*inch])
            equip_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(equip_table)
            elements.append(Spacer(1, 0.2*inch))
        
        # Labor
        await cur.execute("SELECT * FROM labor WHERE room_id = %s", (room['id'],))
        labor = await cur.fetchall()
        if labor:
            elements.append(Paragraph("Labor:", styles['Heading3']))
            labor_data = [['Role', 'Hours', 'Rate', 'Total']]
            for lb in labor:
                total = float(lb['hours']) * float(lb['rate'])
                total_labor += total
                labor_data.append([
                    lb['role_name'],
                    str(lb['hours']),
                    f"${lb['rate']:.2f}/hr",
                    f"${total:.2f}"
                ])
            
            labor_table = Table(labor_data, colWidths=[3*inch, 0.8*inch, 1*inch, 1*inch])
            labor_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(labor_table)
            elements.append(Spacer(1, 0.2*inch))
        
        # Services
        await cur.execute("SELECT * FROM services WHERE room_id = %s", (room['id'],))
        services = await cur.fetchall()
        if services:
            elements.append(Paragraph("Third-Party Services:", styles['Heading3']))
            service_data = [['Service', 'Cost']]
            for sv in services:
                total_services += float(sv['cost'])
                service_data.append([sv['service_name'], f"${sv['cost']:.2f}"])
            
            service_table = Table(service_data, colWidths=[4*inch, 1.5*inch])
            service_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(service_table)
            elements.append(Spacer(1, 0.2*inch))
        
        elements.append(Spacer(1, 0.3*inch))
    
    # Total summary
    grand_total = total_equipment + total_labor + total_services
    summary_data = [
        ['Equipment Total:', f"${total_equipment:.2f}"],
        ['Labor Total:', f"${total_labor:.2f}"],
        ['Services Total:', f"${total_services:.2f}"],
        ['GRAND TOTAL:', f"${grand_total:.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -2), 11),
        ('FONTSIZE', (0, -1), (-1, -1), 14),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(summary_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=quote_{quote_id}.pdf"}
    )

# ============ User Management (Admin) ============

@api_router.get("/users")
async def get_users(user = Depends(require_admin), cur = Depends(get_db)):
    await cur.execute("""
        SELECT u.id, u.username, u.role, u.department_id, u.created_at, d.name as department_name
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        ORDER BY u.created_at DESC
    """)
    return await cur.fetchall()

@api_router.put("/users/{user_id}")
async def update_user(user_id: int, update_data: dict, user = Depends(require_admin), cur = Depends(get_db)):
    fields = []
    values = []
    
    if 'role' in update_data:
        fields.append('role = %s')
        values.append(update_data['role'])
    if 'department_id' in update_data:
        fields.append('department_id = %s')
        values.append(update_data['department_id'])
    
    if fields:
        values.append(user_id)
        await cur.execute(f"UPDATE users SET {', '.join(fields)} WHERE id = %s", tuple(values))
    
    return {"message": "User updated successfully"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: int, user = Depends(require_admin), cur = Depends(get_db)):
    await cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
    return {"message": "User deleted successfully"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    await init_db()
    logger.info("Database initialized")

@app.on_event("shutdown")
async def shutdown():
    global pool
    if pool:
        pool.close()
        await pool.wait_closed()
