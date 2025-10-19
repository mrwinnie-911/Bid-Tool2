from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.responses import StreamingResponse
import aiomysql
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import json
import openpyxl
import io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MySQL connection pool
pool = None

async def get_db_pool():
    global pool
    if pool is None:
        pool = await aiomysql.create_pool(
            host=os.environ.get('MYSQL_HOST', 'localhost'),
            port=int(os.environ.get('MYSQL_PORT', '3306')),
            user=os.environ.get('MYSQL_USER', 'root'),
            password=os.environ.get('MYSQL_PASSWORD', ''),
            db=os.environ.get('MYSQL_DATABASE', 'avlv_quotes'),
            autocommit=True,
            minsize=1,
            maxsize=10
        )
    return pool

async def get_db():
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        async with conn.cursor(aiomysql.DictCursor) as cur:
            yield cur

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# ============ Models ============

class UserCreate(BaseModel):
    username: str
    password: str
    role: str = 'estimator'
    department_id: Optional[int] = None

class UserLogin(BaseModel):
    username: str
    password: str

class DepartmentCreate(BaseModel):
    name: str

class QuoteCreate(BaseModel):
    name: str
    client_name: str
    department_id: int
    company_id: Optional[int] = None  # NEW: Company
    contact_id: Optional[int] = None  # NEW: Point of Contact
    project_address: Optional[str] = None  # NEW: Project address
    description: Optional[str] = None
    equipment_markup_default: float = 20.0  # Default 20% markup
    tax_rate: float = 8.0  # NEW: Default 8%
    tax_enabled: bool = True  # NEW: Enabled by default

class QuoteUpdate(BaseModel):
    name: Optional[str] = None
    client_name: Optional[str] = None
    department_id: Optional[int] = None
    company_id: Optional[int] = None
    contact_id: Optional[int] = None
    project_address: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    equipment_markup_default: Optional[float] = None
    tax_rate: Optional[float] = None
    tax_enabled: Optional[bool] = None

class RoomCreate(BaseModel):
    quote_id: int
    name: str
    quantity: int = 1  # NEW: Room quantity

class SystemCreate(BaseModel):
    room_id: int
    name: str
    description: Optional[str] = None

class EquipmentCreate(BaseModel):
    system_id: int
    item_name: str
    model: Optional[str] = None  # NEW: Model number
    description: Optional[str] = None
    quantity: int
    unit_cost: float  # What you pay
    markup_override: Optional[float] = None  # Override project default
    vendor: Optional[str] = None
    tax_exempt: bool = False  # NEW: Tax exemption flag

class LaborCreate(BaseModel):
    room_id: int
    role_name: str
    cost_rate: float  # What you pay per hour
    sell_rate: float  # What you charge per hour
    hours: float
    department_id: Optional[int] = None

class ServiceCreate(BaseModel):
    room_id: int
    service_name: str
    percentage_of_equipment: float  # Percentage of equipment sell price
    department_id: Optional[int] = None
    description: Optional[str] = None

class TemplateCreate(BaseModel):
    name: str
    department_id: Optional[int] = None
    services: List[Dict[str, Any]]  # List of service templates
    labor: List[Dict[str, Any]]  # List of labor templates
    tax_settings: Dict[str, Any]  # Tax configuration

class VendorPriceCreate(BaseModel):
    item_name: str
    model: Optional[str] = None  # NEW: Model number
    cost: float
    description: Optional[str] = None
    vendor: str
    department_id: Optional[int] = None
    all_departments: bool = False
    expiration_date: Optional[str] = None  # NEW: Price expiration

class CompanyCreate(BaseModel):
    name: str
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None

class ContactCreate(BaseModel):
    company_id: int
    name: str
    title: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None

# ============ Auth Helper Functions ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: int, username: str, role: str) -> str:
    payload = {
        'user_id': user_id,
        'username': username,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(user = Depends(get_current_user)):
    if user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

# ============ Database Initialization ============

async def init_db():
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            # Users table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    username VARCHAR(255) UNIQUE NOT NULL,
                    password_hash VARCHAR(255) NOT NULL,
                    role VARCHAR(50) NOT NULL,
                    department_id INT,
                    azure_enabled BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_username (username),
                    INDEX idx_role (role)
                )
            """)

            # Departments table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS departments (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) UNIQUE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # NEW: Companies table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS companies (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    address TEXT,
                    phone VARCHAR(50),
                    email VARCHAR(255),
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_name (name)
                )
            """)

            # NEW: Contacts table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS contacts (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    company_id INT NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    title VARCHAR(255),
                    phone VARCHAR(50),
                    email VARCHAR(255),
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (company_id) REFERENCES companies(id) ON DELETE CASCADE,
                    INDEX idx_company (company_id)
                )
            """)

            # Quotes table - UPDATED
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS quotes (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    quote_number VARCHAR(50) UNIQUE NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    client_name VARCHAR(255) NOT NULL,
                    department_id INT NOT NULL,
                    company_id INT,
                    contact_id INT,
                    project_address TEXT,
                    description TEXT,
                    status VARCHAR(50) DEFAULT 'draft',
                    version INT DEFAULT 1,
                    equipment_markup_default DECIMAL(5,2) DEFAULT 20.00,
                    tax_rate DECIMAL(5,2) DEFAULT 8.00,
                    tax_enabled BOOLEAN DEFAULT TRUE,
                    created_by INT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    FOREIGN KEY (company_id) REFERENCES companies(id),
                    FOREIGN KEY (contact_id) REFERENCES contacts(id),
                    FOREIGN KEY (created_by) REFERENCES users(id),
                    INDEX idx_status (status),
                    INDEX idx_department (department_id),
                    INDEX idx_created_by (created_by),
                    INDEX idx_quote_number (quote_number)
                )
            """)

            # Quote versions
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS quote_versions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    quote_id INT NOT NULL,
                    version INT NOT NULL,
                    data JSON NOT NULL,
                    changed_by INT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE,
                    FOREIGN KEY (changed_by) REFERENCES users(id),
                    INDEX idx_quote_version (quote_id, version)
                )
            """)

            # Rooms table - UPDATED
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS rooms (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    quote_id INT NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    quantity INT DEFAULT 1,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE,
                    INDEX idx_quote (quote_id)
                )
            """)

            # NEW: Systems table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS systems (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    room_id INT NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
                    INDEX idx_room (room_id)
                )
            """)

            # Equipment table - UPDATED
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS equipment (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    system_id INT NOT NULL,
                    item_name VARCHAR(255) NOT NULL,
                    description TEXT,
                    quantity INT NOT NULL,
                    unit_cost DECIMAL(10, 2) NOT NULL,
                    markup_override DECIMAL(5, 2) NULL,
                    vendor VARCHAR(255),
                    tax_exempt BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (system_id) REFERENCES systems(id) ON DELETE CASCADE,
                    INDEX idx_system (system_id)
                )
            """)

            # Labor table - UPDATED
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS labor (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    room_id INT NOT NULL,
                    role_name VARCHAR(255) NOT NULL,
                    cost_rate DECIMAL(10, 2) NOT NULL,
                    sell_rate DECIMAL(10, 2) NOT NULL,
                    hours DECIMAL(10, 2) NOT NULL,
                    department_id INT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    INDEX idx_room (room_id)
                )
            """)

            # Services table - UPDATED
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS services (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    room_id INT NOT NULL,
                    service_name VARCHAR(255) NOT NULL,
                    percentage_of_equipment DECIMAL(5, 2) NOT NULL,
                    department_id INT,
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    INDEX idx_room (room_id)
                )
            """)

            # Vendor prices - UPDATED
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS vendor_prices (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    item_name VARCHAR(255) NOT NULL,
                    cost DECIMAL(10, 2) NOT NULL,
                    description TEXT,
                    vendor VARCHAR(255) NOT NULL,
                    department_id INT,
                    all_departments BOOLEAN DEFAULT FALSE,
                    imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    INDEX idx_vendor (vendor),
                    INDEX idx_item (item_name)
                )
            """)

            # NEW: Templates table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS templates (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    department_id INT,
                    services_json JSON NOT NULL,
                    labor_json JSON NOT NULL,
                    tax_settings_json JSON NOT NULL,
                    created_by INT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (department_id) REFERENCES departments(id),
                    FOREIGN KEY (created_by) REFERENCES users(id),
                    INDEX idx_department (department_id)
                )
            """)

            # Metrics table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS metrics (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    metric_name VARCHAR(255) NOT NULL,
                    metric_type VARCHAR(100) NOT NULL,
                    config JSON NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                    INDEX idx_user (user_id)
                )
            """)

            # Approvals table
            await cur.execute("""
                CREATE TABLE IF NOT EXISTS approvals (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    quote_id INT NOT NULL,
                    approver_id INT NOT NULL,
                    status VARCHAR(50) DEFAULT 'pending',
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE,
                    FOREIGN KEY (approver_id) REFERENCES users(id),
                    INDEX idx_quote (quote_id),
                    INDEX idx_status (status)
                )
            """)

            # Create default admin user if not exists
            await cur.execute("SELECT id FROM users WHERE username = 'admin'")
            if not await cur.fetchone():
                admin_password = hash_password('admin123')
                await cur.execute(
                    "INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s)",
                    ('admin', admin_password, 'admin')
                )

            # Create default departments if not exist
            for dept_name in ['AV', 'LV', 'IT']:
                await cur.execute("SELECT id FROM departments WHERE name = %s", (dept_name,))
                if not await cur.fetchone():
                    await cur.execute("INSERT INTO departments (name) VALUES (%s)", (dept_name,))

print("Database initialization script created successfully!")
# ============ Auth Routes ============

@api_router.post("/auth/register")
async def register(user_data: UserCreate, current_user = Depends(require_admin), cur = Depends(get_db)):
    """Admin only - register new users"""
    password_hash = hash_password(user_data.password)
    try:
        await cur.execute(
            "INSERT INTO users (username, password_hash, role, department_id) VALUES (%s, %s, %s, %s)",
            (user_data.username, password_hash, user_data.role, user_data.department_id)
        )
        return {"message": "User created successfully"}
    except aiomysql.IntegrityError:
        raise HTTPException(status_code=400, detail="Username already exists")

@api_router.post("/auth/login")
async def login(credentials: UserLogin, cur = Depends(get_db)):
    await cur.execute("SELECT * FROM users WHERE username = %s", (credentials.username,))
    user = await cur.fetchone()
    
    if not user or not verify_password(credentials.password, user['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user['id'], user['username'], user['role'])
    return {
        "token": token,
        "user": {
            "id": user['id'],
            "username": user['username'],
            "role": user['role'],
            "department_id": user['department_id']
        }
    }

@api_router.get("/auth/me")
async def get_me(user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT id, username, role, department_id, created_at FROM users WHERE id = %s", (user['user_id'],))
    user_data = await cur.fetchone()
    if not user_data:
        raise HTTPException(status_code=404, detail="User not found")
    return user_data

# ============ Department Routes ============

@api_router.post("/departments")
async def create_department(dept: DepartmentCreate, user = Depends(require_admin), cur = Depends(get_db)):
    try:
        await cur.execute("INSERT INTO departments (name) VALUES (%s)", (dept.name,))
        return {"message": "Department created successfully"}
    except aiomysql.IntegrityError:
        raise HTTPException(status_code=400, detail="Department already exists")

@api_router.get("/departments")
async def get_departments(user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM departments ORDER BY name")
    return await cur.fetchall()

@api_router.put("/departments/{dept_id}")
async def update_department(dept_id: int, dept: DepartmentCreate, user = Depends(require_admin), cur = Depends(get_db)):
    await cur.execute("UPDATE departments SET name = %s WHERE id = %s", (dept.name, dept_id))
    return {"message": "Department updated successfully"}

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: int, user = Depends(require_admin), cur = Depends(get_db)):
    await cur.execute("DELETE FROM departments WHERE id = %s", (dept_id,))
    return {"message": "Department deleted successfully"}

# ============ Quote Routes ============

@api_router.post("/quotes")
async def create_quote(quote: QuoteCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        """INSERT INTO quotes (name, client_name, department_id, description, equipment_markup_default, 
           tax_rate, tax_enabled, created_by) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
        (quote.name, quote.client_name, quote.department_id, quote.description, 
         quote.equipment_markup_default, quote.tax_rate, quote.tax_enabled, user['user_id'])
    )
    quote_id = cur.lastrowid
    return {"id": quote_id, "message": "Quote created successfully"}

@api_router.get("/quotes")
async def get_quotes(user = Depends(get_current_user), cur = Depends(get_db)):
    if user['role'] == 'admin':
        await cur.execute("""
            SELECT q.*, d.name as department_name, u.username as created_by_username
            FROM quotes q
            LEFT JOIN departments d ON q.department_id = d.id
            LEFT JOIN users u ON q.created_by = u.id
            ORDER BY q.updated_at DESC
        """)
    else:
        await cur.execute("""
            SELECT q.*, d.name as department_name, u.username as created_by_username
            FROM quotes q
            LEFT JOIN departments d ON q.department_id = d.id
            LEFT JOIN users u ON q.created_by = u.id
            WHERE q.department_id = (SELECT department_id FROM users WHERE id = %s)
            ORDER BY q.updated_at DESC
        """, (user['user_id'],))
    
    return await cur.fetchall()

@api_router.get("/quotes/{quote_id}")
async def get_quote(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("""
        SELECT q.*, d.name as department_name, u.username as created_by_username
        FROM quotes q
        LEFT JOIN departments d ON q.department_id = d.id
        LEFT JOIN users u ON q.created_by = u.id
        WHERE q.id = %s
    """, (quote_id,))
    quote = await cur.fetchone()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    return quote

@api_router.put("/quotes/{quote_id}")
async def update_quote(quote_id: int, quote_data: QuoteUpdate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT version FROM quotes WHERE id = %s", (quote_id,))
    current = await cur.fetchone()
    if not current:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Save version
    await cur.execute("SELECT * FROM quotes WHERE id = %s", (quote_id,))
    quote_snapshot = await cur.fetchone()
    await cur.execute(
        "INSERT INTO quote_versions (quote_id, version, data, changed_by) VALUES (%s, %s, %s, %s)",
        (quote_id, current['version'], json.dumps(quote_snapshot, default=str), user['user_id'])
    )
    
    # Update quote
    update_fields = []
    values = []
    for field, value in quote_data.model_dump(exclude_unset=True).items():
        update_fields.append(f"{field} = %s")
        values.append(value)
    
    if update_fields:
        update_fields.append("version = version + 1")
        values.append(quote_id)
        await cur.execute(
            f"UPDATE quotes SET {', '.join(update_fields)} WHERE id = %s",
            tuple(values)
        )
    
    return {"message": "Quote updated successfully"}

@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM quotes WHERE id = %s", (quote_id,))
    return {"message": "Quote deleted successfully"}

# ============ Room Routes ============

@api_router.post("/rooms")
async def create_room(room: RoomCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO rooms (quote_id, name, quantity) VALUES (%s, %s, %s)",
        (room.quote_id, room.name, room.quantity)
    )
    return {"id": cur.lastrowid, "message": "Room created successfully"}

@api_router.get("/rooms/quote/{quote_id}")
async def get_rooms_by_quote(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM rooms WHERE quote_id = %s", (quote_id,))
    return await cur.fetchall()

@api_router.put("/rooms/{room_id}")
async def update_room(room_id: int, room: RoomCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "UPDATE rooms SET name = %s, quantity = %s WHERE id = %s",
        (room.name, room.quantity, room_id)
    )
    return {"message": "Room updated successfully"}

@api_router.delete("/rooms/{room_id}")
async def delete_room(room_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM rooms WHERE id = %s", (room_id,))
    return {"message": "Room deleted successfully"}

# ============ System Routes (NEW) ============

@api_router.post("/systems")
async def create_system(system: SystemCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "INSERT INTO systems (room_id, name, description) VALUES (%s, %s, %s)",
        (system.room_id, system.name, system.description)
    )
    return {"id": cur.lastrowid, "message": "System created successfully"}

@api_router.get("/systems/room/{room_id}")
async def get_systems_by_room(room_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM systems WHERE room_id = %s", (room_id,))
    return await cur.fetchall()

@api_router.put("/systems/{system_id}")
async def update_system(system_id: int, system: SystemCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "UPDATE systems SET name = %s, description = %s WHERE id = %s",
        (system.name, system.description, system_id)
    )
    return {"message": "System updated successfully"}

@api_router.delete("/systems/{system_id}")
async def delete_system(system_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM systems WHERE id = %s", (system_id,))
    return {"message": "System deleted successfully"}

# ============ Equipment Routes (UPDATED) ============

@api_router.post("/equipment")
async def create_equipment(equip: EquipmentCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        """INSERT INTO equipment (system_id, item_name, description, quantity, unit_cost, 
           markup_override, vendor, tax_exempt) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
        (equip.system_id, equip.item_name, equip.description, equip.quantity, 
         equip.unit_cost, equip.markup_override, equip.vendor, equip.tax_exempt)
    )
    return {"id": cur.lastrowid, "message": "Equipment created successfully"}

@api_router.get("/equipment/system/{system_id}")
async def get_equipment_by_system(system_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    # Get quote's default markup
    await cur.execute("""
        SELECT q.equipment_markup_default 
        FROM quotes q
        JOIN rooms r ON q.id = r.quote_id
        JOIN systems s ON r.id = s.room_id
        WHERE s.id = %s
    """, (system_id,))
    quote_data = await cur.fetchone()
    default_markup = float(quote_data['equipment_markup_default']) if quote_data else 20.0
    
    await cur.execute("SELECT * FROM equipment WHERE system_id = %s", (system_id,))
    equipment = await cur.fetchall()
    
    # Calculate prices with markup
    for eq in equipment:
        markup = float(eq['markup_override']) if eq['markup_override'] else default_markup
        unit_cost = float(eq['unit_cost'])
        unit_price = unit_cost * (1 + markup / 100)
        eq['unit_price'] = round(unit_price, 2)
        eq['total_cost'] = round(unit_cost * eq['quantity'], 2)
        eq['total_price'] = round(unit_price * eq['quantity'], 2)
        eq['margin_dollars'] = round((unit_price - unit_cost) * eq['quantity'], 2)
        eq['margin_percent'] = round(markup, 2)
    
    return equipment

@api_router.put("/equipment/{equip_id}")
async def update_equipment(equip_id: int, equip: EquipmentCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        """UPDATE equipment SET item_name = %s, description = %s, quantity = %s, unit_cost = %s,
           markup_override = %s, vendor = %s, tax_exempt = %s WHERE id = %s""",
        (equip.item_name, equip.description, equip.quantity, equip.unit_cost,
         equip.markup_override, equip.vendor, equip.tax_exempt, equip_id)
    )
    return {"message": "Equipment updated successfully"}

@api_router.delete("/equipment/{equip_id}")
async def delete_equipment(equip_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM equipment WHERE id = %s", (equip_id,))
    return {"message": "Equipment deleted successfully"}

# ============ Labor Routes (UPDATED) ============

@api_router.post("/labor")
async def create_labor(labor: LaborCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        """INSERT INTO labor (room_id, role_name, cost_rate, sell_rate, hours, department_id) 
           VALUES (%s, %s, %s, %s, %s, %s)""",
        (labor.room_id, labor.role_name, labor.cost_rate, labor.sell_rate, labor.hours, labor.department_id)
    )
    return {"id": cur.lastrowid, "message": "Labor created successfully"}

@api_router.get("/labor/room/{room_id}")
async def get_labor_by_room(room_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM labor WHERE room_id = %s", (room_id,))
    labor = await cur.fetchall()
    
    # Calculate totals and margins
    for lb in labor:
        lb['total_cost'] = round(float(lb['cost_rate']) * float(lb['hours']), 2)
        lb['total_price'] = round(float(lb['sell_rate']) * float(lb['hours']), 2)
        lb['margin_dollars'] = round(lb['total_price'] - lb['total_cost'], 2)
        lb['margin_percent'] = round((lb['margin_dollars'] / lb['total_cost'] * 100) if lb['total_cost'] > 0 else 0, 2)
    
    return labor

@api_router.put("/labor/{labor_id}")
async def update_labor(labor_id: int, labor: LaborCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        """UPDATE labor SET role_name = %s, cost_rate = %s, sell_rate = %s, hours = %s,
           department_id = %s WHERE id = %s""",
        (labor.role_name, labor.cost_rate, labor.sell_rate, labor.hours, labor.department_id, labor_id)
    )
    return {"message": "Labor updated successfully"}

@api_router.delete("/labor/{labor_id}")
async def delete_labor(labor_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM labor WHERE id = %s", (labor_id,))
    return {"message": "Labor deleted successfully"}

# ============ Service Routes (UPDATED) ============

@api_router.post("/services")
async def create_service(service: ServiceCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        """INSERT INTO services (room_id, service_name, percentage_of_equipment, department_id, description) 
           VALUES (%s, %s, %s, %s, %s)""",
        (service.room_id, service.service_name, service.percentage_of_equipment, service.department_id, service.description)
    )
    return {"id": cur.lastrowid, "message": "Service created successfully"}

@api_router.get("/services/room/{room_id}")
async def get_services_by_room(room_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM services WHERE room_id = %s", (room_id,))
    return await cur.fetchall()

@api_router.put("/services/{service_id}")
async def update_service(service_id: int, service: ServiceCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        """UPDATE services SET service_name = %s, percentage_of_equipment = %s, 
           department_id = %s, description = %s WHERE id = %s""",
        (service.service_name, service.percentage_of_equipment, service.department_id, service.description, service_id)
    )
    return {"message": "Service updated successfully"}

@api_router.delete("/services/{service_id}")
async def delete_service(service_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM services WHERE id = %s", (service_id,))
    return {"message": "Service deleted successfully"}


# ============ Template Routes (NEW) ============

@api_router.post("/templates")
async def create_template(template: TemplateCreate, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        """INSERT INTO templates (name, department_id, services_json, labor_json, tax_settings_json, created_by)
           VALUES (%s, %s, %s, %s, %s, %s)""",
        (template.name, template.department_id, json.dumps(template.services), 
         json.dumps(template.labor), json.dumps(template.tax_settings), user['user_id'])
    )
    return {"id": cur.lastrowid, "message": "Template created successfully"}

@api_router.get("/templates")
async def get_templates(user = Depends(get_current_user), cur = Depends(get_db)):
    if user['role'] == 'admin':
        await cur.execute("SELECT * FROM templates ORDER BY created_at DESC")
    else:
        await cur.execute(
            """SELECT * FROM templates WHERE department_id IS NULL 
               OR department_id = (SELECT department_id FROM users WHERE id = %s)
               ORDER BY created_at DESC""",
            (user['user_id'],)
        )
    templates = await cur.fetchall()
    for t in templates:
        t['services'] = json.loads(t['services_json'])
        t['labor'] = json.loads(t['labor_json'])
        t['tax_settings'] = json.loads(t['tax_settings_json'])
    return templates

@api_router.get("/templates/{template_id}")
async def get_template(template_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM templates WHERE id = %s", (template_id,))
    template = await cur.fetchone()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    template['services'] = json.loads(template['services_json'])
    template['labor'] = json.loads(template['labor_json'])
    template['tax_settings'] = json.loads(template['tax_settings_json'])
    return template

@api_router.delete("/templates/{template_id}")
async def delete_template(template_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("DELETE FROM templates WHERE id = %s", (template_id,))
    return {"message": "Template deleted successfully"}

@api_router.post("/quotes/{quote_id}/apply-template/{template_id}")
async def apply_template(quote_id: int, template_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    # Get template
    await cur.execute("SELECT * FROM templates WHERE id = %s", (template_id,))
    template = await cur.fetchone()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    services = json.loads(template['services_json'])
    labor = json.loads(template['labor_json'])
    tax_settings = json.loads(template['tax_settings_json'])
    
    # Update quote tax settings
    await cur.execute(
        "UPDATE quotes SET tax_rate = %s, tax_enabled = %s WHERE id = %s",
        (tax_settings.get('tax_rate', 0), tax_settings.get('tax_enabled', False), quote_id)
    )
    
    return {
        "message": "Template applied",
        "services": services,
        "labor": labor,
        "tax_settings": tax_settings
    }

# ============ Financial Breakdown Route (NEW) ============

@api_router.get("/quotes/{quote_id}/financials")
async def get_quote_financials(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    # Get quote
    await cur.execute("SELECT * FROM quotes WHERE id = %s", (quote_id,))
    quote = await cur.fetchone()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    financials = {
        "rooms": [],
        "totals": {
            "equipment_cost": 0,
            "equipment_price": 0,
            "equipment_margin": 0,
            "labor_cost": 0,
            "labor_price": 0,
            "labor_margin": 0,
            "services_cost": 0,
            "services_price": 0,
            "subtotal": 0,
            "tax": 0,
            "grand_total": 0,
            "total_margin": 0,
            "margin_percent": 0
        }
    }
    
    # Get rooms
    await cur.execute("SELECT * FROM rooms WHERE quote_id = %s", (quote_id,))
    rooms = await cur.fetchall()
    
    for room in rooms:
        room_data = {
            "id": room['id'],
            "name": room['name'],
            "quantity": room['quantity'],
            "systems": [],
            "labor": [],
            "services": [],
            "equipment_cost": 0,
            "equipment_price": 0,
            "labor_cost": 0,
            "labor_price": 0,
            "services_price": 0,
            "subtotal": 0,
            "margin": 0
        }
        
        # Get systems
        await cur.execute("SELECT * FROM systems WHERE room_id = %s", (room['id'],))
        systems = await cur.fetchall()
        
        for system in systems:
            system_data = {
                "id": system['id'],
                "name": system['name'],
                "equipment": [],
                "equipment_cost": 0,
                "equipment_price": 0,
                "margin": 0
            }
            
            # Get equipment
            await cur.execute("SELECT * FROM equipment WHERE system_id = %s", (system['id'],))
            equipment = await cur.fetchall()
            
            for eq in equipment:
                markup = float(eq['markup_override']) if eq['markup_override'] else float(quote['equipment_markup_default'])
                unit_cost = float(eq['unit_cost'])
                unit_price = unit_cost * (1 + markup / 100)
                total_cost = unit_cost * eq['quantity']
                total_price = unit_price * eq['quantity']
                
                system_data['equipment'].append({
                    "id": eq['id'],
                    "item_name": eq['item_name'],
                    "quantity": eq['quantity'],
                    "unit_cost": round(unit_cost, 2),
                    "unit_price": round(unit_price, 2),
                    "total_cost": round(total_cost, 2),
                    "total_price": round(total_price, 2),
                    "margin": round(total_price - total_cost, 2),
                    "markup_percent": round(markup, 2),
                    "tax_exempt": eq['tax_exempt']
                })
                
                system_data['equipment_cost'] += total_cost
                system_data['equipment_price'] += total_price
            
            system_data['equipment_cost'] = round(system_data['equipment_cost'], 2)
            system_data['equipment_price'] = round(system_data['equipment_price'], 2)
            system_data['margin'] = round(system_data['equipment_price'] - system_data['equipment_cost'], 2)
            
            room_data['systems'].append(system_data)
            room_data['equipment_cost'] += system_data['equipment_cost']
            room_data['equipment_price'] += system_data['equipment_price']
        
        # Get labor
        await cur.execute("SELECT * FROM labor WHERE room_id = %s", (room['id'],))
        labor = await cur.fetchall()
        
        for lb in labor:
            total_cost = float(lb['cost_rate']) * float(lb['hours'])
            total_price = float(lb['sell_rate']) * float(lb['hours'])
            
            room_data['labor'].append({
                "id": lb['id'],
                "role_name": lb['role_name'],
                "cost_rate": float(lb['cost_rate']),
                "sell_rate": float(lb['sell_rate']),
                "hours": float(lb['hours']),
                "total_cost": round(total_cost, 2),
                "total_price": round(total_price, 2),
                "margin": round(total_price - total_cost, 2)
            })
            
            room_data['labor_cost'] += total_cost
            room_data['labor_price'] += total_price
        
        # Get services (calculated as % of equipment)
        await cur.execute("SELECT * FROM services WHERE room_id = %s", (room['id'],))
        services = await cur.fetchall()
        
        for svc in services:
            percentage = float(svc['percentage_of_equipment'])
            service_price = room_data['equipment_price'] * (percentage / 100)
            
            room_data['services'].append({
                "id": svc['id'],
                "service_name": svc['service_name'],
                "percentage": percentage,
                "calculated_price": round(service_price, 2)
            })
            
            room_data['services_price'] += service_price
        
        # Multiply by room quantity
        room_data['equipment_cost'] = round(room_data['equipment_cost'] * room['quantity'], 2)
        room_data['equipment_price'] = round(room_data['equipment_price'] * room['quantity'], 2)
        room_data['labor_cost'] = round(room_data['labor_cost'] * room['quantity'], 2)
        room_data['labor_price'] = round(room_data['labor_price'] * room['quantity'], 2)
        room_data['services_price'] = round(room_data['services_price'] * room['quantity'], 2)
        room_data['subtotal'] = round(room_data['equipment_price'] + room_data['labor_price'] + room_data['services_price'], 2)
        room_data['margin'] = round((room_data['equipment_price'] - room_data['equipment_cost']) + (room_data['labor_price'] - room_data['labor_cost']) + room_data['services_price'], 2)
        
        financials['rooms'].append(room_data)
        
        # Add to totals
        financials['totals']['equipment_cost'] += room_data['equipment_cost']
        financials['totals']['equipment_price'] += room_data['equipment_price']
        financials['totals']['labor_cost'] += room_data['labor_cost']
        financials['totals']['labor_price'] += room_data['labor_price']
        financials['totals']['services_price'] += room_data['services_price']
    
    # Calculate totals
    financials['totals']['equipment_cost'] = round(financials['totals']['equipment_cost'], 2)
    financials['totals']['equipment_price'] = round(financials['totals']['equipment_price'], 2)
    financials['totals']['equipment_margin'] = round(financials['totals']['equipment_price'] - financials['totals']['equipment_cost'], 2)
    
    financials['totals']['labor_cost'] = round(financials['totals']['labor_cost'], 2)
    financials['totals']['labor_price'] = round(financials['totals']['labor_price'], 2)
    financials['totals']['labor_margin'] = round(financials['totals']['labor_price'] - financials['totals']['labor_cost'], 2)
    
    financials['totals']['services_price'] = round(financials['totals']['services_price'], 2)
    financials['totals']['services_cost'] = 0  # Services are pure margin
    
    financials['totals']['subtotal'] = round(
        financials['totals']['equipment_price'] + 
        financials['totals']['labor_price'] + 
        financials['totals']['services_price'], 2
    )
    
    # Calculate tax (only on non-exempt equipment)
    if quote['tax_enabled']:
        tax_rate = float(quote['tax_rate'])
        # For now, tax all equipment. In future, filter by tax_exempt flag
        taxable_amount = financials['totals']['equipment_price']
        financials['totals']['tax'] = round(taxable_amount * (tax_rate / 100), 2)
    
    financials['totals']['grand_total'] = round(financials['totals']['subtotal'] + financials['totals']['tax'], 2)
    
    financials['totals']['total_margin'] = round(
        financials['totals']['equipment_margin'] + 
        financials['totals']['labor_margin'] + 
        financials['totals']['services_price'], 2
    )
    
    total_cost = financials['totals']['equipment_cost'] + financials['totals']['labor_cost']
    if total_cost > 0:
        financials['totals']['margin_percent'] = round((financials['totals']['total_margin'] / total_cost) * 100, 2)
    
    return financials

# ============ Vendor Prices ============

@api_router.post("/vendor-prices/import")
async def import_vendor_prices(file: UploadFile = File(...), user = Depends(get_current_user), cur = Depends(get_db)):
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active
    
    headers = [cell.value for cell in sheet[1]]
    return {"headers": headers, "row_count": sheet.max_row - 1}

@api_router.post("/vendor-prices/import-mapped")
async def import_vendor_prices_mapped(
    file: UploadFile = File(...),
    mapping: str = Form(...),
    vendor: str = Form(...),
    department_id: Optional[str] = Form(None),
    all_departments: str = Form("false"),
    user = Depends(get_current_user),
    cur = Depends(get_db)
):
    try:
        mapping_dict = json.loads(mapping)
        all_depts = all_departments.lower() == "true"
        dept_id = int(department_id) if department_id and department_id != "null" else None
        
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        imported_count = 0
        errors = []
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Get mapped columns
                item_name = None
                cost = None
                description = None
                
                if mapping_dict.get('item_name') and mapping_dict['item_name'] in headers:
                    item_name = row[headers.index(mapping_dict['item_name'])]
                
                if mapping_dict.get('price') and mapping_dict['price'] in headers:
                    cost = row[headers.index(mapping_dict['price'])]
                
                if mapping_dict.get('description') and mapping_dict['description'] in headers:
                    description = row[headers.index(mapping_dict['description'])]
                
                if item_name and cost:
                    await cur.execute(
                        """INSERT INTO vendor_prices (item_name, cost, description, vendor, department_id, all_departments) 
                           VALUES (%s, %s, %s, %s, %s, %s)""",
                        (str(item_name), float(cost), str(description) if description else None, 
                         vendor, dept_id, all_depts)
                    )
                    imported_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                logging.error(f"Error importing row {row_num}: {e}")
        
        result = {"message": f"Imported {imported_count} items successfully"}
        if errors:
            result['errors'] = errors[:10]  # Return first 10 errors
        
        return result
    except Exception as e:
        logging.error(f"Import error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

@api_router.get("/vendor-prices")
async def get_vendor_prices(user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute("SELECT * FROM vendor_prices ORDER BY imported_at DESC LIMIT 100")
    return await cur.fetchall()

@api_router.get("/vendor-prices/search")
async def search_vendor_prices(q: str, user = Depends(get_current_user), cur = Depends(get_db)):
    await cur.execute(
        "SELECT * FROM vendor_prices WHERE item_name LIKE %s OR description LIKE %s LIMIT 50",
        (f"%{q}%", f"%{q}%")
    )
    return await cur.fetchall()

# ============ Dashboard ============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user = Depends(get_current_user), cur = Depends(get_db)):
    stats = {}
    
    dept_filter = "" if user['role'] == 'admin' else f"WHERE department_id = (SELECT department_id FROM users WHERE id = {user['user_id']})"
    
    await cur.execute(f"""
        SELECT d.name, COUNT(q.id) as count
        FROM departments d
        LEFT JOIN quotes q ON d.id = q.department_id
        {dept_filter if dept_filter else ""}
        GROUP BY d.id, d.name
    """)
    stats['quotes_by_department'] = await cur.fetchall()
    
    await cur.execute(f"SELECT status, COUNT(*) as count FROM quotes {dept_filter} GROUP BY status")
    stats['quotes_by_status'] = await cur.fetchall()
    
    await cur.execute(f"""
        SELECT DATE(created_at) as date, COUNT(*) as count
        FROM quotes
        WHERE created_at >= DATE_SUB(NOW(), INTERVAL 90 DAY)
        {dept_filter}
        GROUP BY DATE(created_at)
        ORDER BY date
    """)
    stats['recent_quotes'] = await cur.fetchall()
    
    return stats

# ============ User Management ============

@api_router.get("/users")
async def get_users(user = Depends(require_admin), cur = Depends(get_db)):
    await cur.execute("""
        SELECT u.id, u.username, u.role, u.department_id, u.created_at, d.name as department_name
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        ORDER BY u.created_at DESC
    """)
    return await cur.fetchall()

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: int, user = Depends(require_admin), cur = Depends(get_db)):
    await cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
    return {"message": "User deleted successfully"}

# ============ Enhanced User Management ============

@api_router.put("/users/{user_id}")
async def update_user(user_id: int, update_data: dict, user = Depends(require_admin), cur = Depends(get_db)):
    """Admin - Update user role and department"""
    fields = []
    values = []
    
    if 'role' in update_data:
        fields.append('role = %s')
        values.append(update_data['role'])
    if 'department_id' in update_data:
        fields.append('department_id = %s')
        values.append(update_data['department_id'])
    
    if fields:
        values.append(user_id)
        await cur.execute(f"UPDATE users SET {', '.join(fields)} WHERE id = %s", tuple(values))
    
    return {"message": "User updated successfully"}

# ============ Quote Version Management ============

@api_router.get("/quotes/{quote_id}/versions")
async def get_quote_versions(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    """Get all versions of a quote"""
    await cur.execute("""
        SELECT qv.id, qv.version, qv.created_at, u.username as changed_by_username,
               qv.data
        FROM quote_versions qv
        LEFT JOIN users u ON qv.changed_by = u.id
        WHERE qv.quote_id = %s
        ORDER BY qv.version DESC
    """, (quote_id,))
    versions = await cur.fetchall()
    
    # Parse JSON data
    for v in versions:
        if isinstance(v['data'], str):
            v['data'] = json.loads(v['data'])
    
    return versions

@api_router.post("/quotes/{quote_id}/restore-version/{version}")
async def restore_quote_version(quote_id: int, version: int, user = Depends(get_current_user), cur = Depends(get_db)):
    """Restore a quote to a specific version"""
    # Get the version data
    await cur.execute(
        "SELECT data FROM quote_versions WHERE quote_id = %s AND version = %s",
        (quote_id, version)
    )
    version_data = await cur.fetchone()
    if not version_data:
        raise HTTPException(status_code=404, detail="Version not found")
    
    # Save current state as new version
    await cur.execute("SELECT * FROM quotes WHERE id = %s", (quote_id,))
    current = await cur.fetchone()
    await cur.execute(
        "INSERT INTO quote_versions (quote_id, version, data, changed_by) VALUES (%s, %s, %s, %s)",
        (quote_id, current['version'], json.dumps(current, default=str), user['user_id'])
    )
    
    # Restore the old version
    old_data = json.loads(version_data['data']) if isinstance(version_data['data'], str) else version_data['data']
    
    await cur.execute("""
        UPDATE quotes SET 
        name = %s, client_name = %s, description = %s, 
        equipment_markup_default = %s, tax_rate = %s, tax_enabled = %s,
        version = version + 1
        WHERE id = %s
    """, (
        old_data['name'], old_data['client_name'], old_data.get('description'),
        old_data.get('equipment_markup_default', 20.0), 
        old_data.get('tax_rate', 0.0), 
        old_data.get('tax_enabled', False),
        quote_id
    ))
    
    return {"message": f"Quote restored to version {version}"}

# ============ Quote Status Management ============

@api_router.put("/quotes/{quote_id}/status")
async def update_quote_status(quote_id: int, status: str, user = Depends(get_current_user), cur = Depends(get_db)):
    """Update quote status"""
    valid_statuses = ['draft', 'pending', 'approved', 'rejected', 'revision']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
    
    await cur.execute("UPDATE quotes SET status = %s WHERE id = %s", (status, quote_id))
    return {"message": "Quote status updated successfully", "status": status}

# ============ BOM Generation ============

@api_router.get("/quotes/{quote_id}/bom")
async def generate_bom(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    """Generate Bill of Materials"""
    # Get quote
    await cur.execute("SELECT * FROM quotes WHERE id = %s", (quote_id,))
    quote = await cur.fetchone()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    bom = {
        "quote_id": quote_id,
        "quote_name": quote['name'],
        "client_name": quote['client_name'],
        "status": quote['status'],
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "items": []
    }
    
    # Get all rooms
    await cur.execute("SELECT * FROM rooms WHERE quote_id = %s", (quote_id,))
    rooms = await cur.fetchall()
    
    equipment_summary = {}  # Group by item for BOM
    
    for room in rooms:
        room_qty = room['quantity']
        
        # Get systems
        await cur.execute("SELECT * FROM systems WHERE room_id = %s", (room['id'],))
        systems = await cur.fetchall()
        
        for system in systems:
            # Get equipment
            await cur.execute("SELECT * FROM equipment WHERE system_id = %s", (system['id'],))
            equipment = await cur.fetchall()
            
            for eq in equipment:
                total_qty = eq['quantity'] * room_qty
                key = f"{eq['item_name']}|{eq['vendor'] or 'N/A'}"
                
                if key in equipment_summary:
                    equipment_summary[key]['quantity'] += total_qty
                else:
                    equipment_summary[key] = {
                        "item_name": eq['item_name'],
                        "description": eq['description'],
                        "vendor": eq['vendor'],
                        "quantity": total_qty,
                        "unit_cost": float(eq['unit_cost']),
                        "locations": []
                    }
                
                equipment_summary[key]['locations'].append({
                    "room": room['name'],
                    "system": system['name'],
                    "quantity": eq['quantity'],
                    "room_quantity": room_qty
                })
    
    # Convert to list
    bom['items'] = list(equipment_summary.values())
    
    # Calculate totals
    bom['total_items'] = len(bom['items'])
    bom['total_quantity'] = sum(item['quantity'] for item in bom['items'])
    bom['total_cost'] = round(sum(item['quantity'] * item['unit_cost'] for item in bom['items']), 2)
    
    return bom

@api_router.get("/quotes/{quote_id}/bom/export")
async def export_bom_excel(quote_id: int, user = Depends(get_current_user), cur = Depends(get_db)):
    """Export BOM as Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    # Get BOM data
    bom_data = await generate_bom(quote_id, user, cur)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill of Materials"
    
    # Header
    ws['A1'] = f"Bill of Materials - {bom_data['quote_name']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Client: {bom_data['client_name']}"
    ws['A3'] = f"Generated: {bom_data['generated_at']}"
    
    # Column headers
    headers = ['Item Name', 'Description', 'Vendor', 'Total Quantity', 'Unit Cost', 'Total Cost', 'Locations']
    header_row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Data rows
    row = header_row + 1
    for item in bom_data['items']:
        ws.cell(row=row, column=1).value = item['item_name']
        ws.cell(row=row, column=2).value = item['description']
        ws.cell(row=row, column=3).value = item['vendor']
        ws.cell(row=row, column=4).value = item['quantity']
        ws.cell(row=row, column=5).value = item['unit_cost']
        ws.cell(row=row, column=6).value = item['quantity'] * item['unit_cost']
        
        # Locations
        locations_text = "; ".join([
            f"{loc['room']} ({loc['system']}): {loc['quantity']}×{loc['room_quantity']}"
            for loc in item['locations']
        ])
        ws.cell(row=row, column=7).value = locations_text
        row += 1
    
    # Totals
    row += 1
    ws.cell(row=row, column=1).value = "TOTALS"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=4).value = bom_data['total_quantity']
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=6).value = bom_data['total_cost']
    ws.cell(row=row, column=6).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 50
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=BOM_Quote_{quote_id}.xlsx"}
    )

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    await init_db()
    logger.info("Database initialized with enhanced schema")

@app.on_event("shutdown")
async def shutdown():
    global pool
    if pool:
        pool.close()
        await pool.wait_closed()
